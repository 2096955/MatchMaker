"""Build JPMC_SCUDO_POAP_TCO_v1.xlsx from scratch with openpyxl.

Contract: JPMC_TCO_design.md + SEGRO_structure_spec.md (mechanisms only).
USD / us-east-1 only. Every computed cell is an Excel formula.
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule  # noqa: F401 -- used in build_gantt
from openpyxl.chart import BarChart, Reference

HERE = os.path.dirname(os.path.abspath(__file__))
RATES_JSON = os.path.join(HERE, "research", "rates.json")

# ---- shared styles ----
MONEY_FMT = "$#,##0;($#,##0);-"
MONEY2_FMT = "$#,##0.00;($#,##0.00);-"
MONEY4_FMT = "$#,##0.0000;($#,##0.0000);-"
MONEY8_FMT = "$#,##0.00000000;($#,##0.00000000);-"
PCT_FMT = "0.0%"
QTY_FMT = "#,##0;(#,##0);-"
HIDE_FMT = ";;;"

BLUE = Font(color="FF0000FF")
BLACK = Font(color="FF000000")
GREEN = Font(color="FF008000")
RED_BOLD = Font(color="FFCC171E", bold=True)
GREY = Font(color="FF808080")
WHITE_BOLD_14 = Font(color="FFFFFFFF", bold=True, size=14)
WHITE_BOLD = Font(color="FFFFFFFF", bold=True)

TITLE_FILL = PatternFill("solid", fgColor="FF1F1F1F")
SECTION_FILL = PatternFill("solid", fgColor="FF595959")
HEADER_FILL = PatternFill("solid", fgColor="FFF2F2F2")
SUBTOTAL_FILL = PatternFill("solid", fgColor="FFE2EFDA")
AMBER_FILL = PatternFill("solid", fgColor="FFFFF2CC")
BAR_FILL = PatternFill("solid", fgColor="FF2E7D32")
BAR_FILL_CHILD = PatternFill("solid", fgColor="FFA1C5A3")

THIN_BOTTOM = Border(bottom=Side(style="thin"))


def style_title(ws, last_col_letter, text, height=18):
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = text
    ws["A1"].font = WHITE_BOLD_14
    ws["A1"].fill = TITLE_FILL
    ws.row_dimensions[1].height = height


def style_banner(ws, last_col_letter, formula, height=30):
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = formula
    ws["A2"].font = Font(italic=True, color="FF808080", size=11)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = height


def style_section(ws, row, last_col_letter, text):
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    ws[f"A{row}"] = text
    ws[f"A{row}"].font = WHITE_BOLD
    ws[f"A{row}"].fill = SECTION_FILL


def style_header_row(ws, row, last_col):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = THIN_BOTTOM


def status_cell(ws, cell_ref, status):
    ws[cell_ref] = status
    ws[cell_ref].font = RED_BOLD


def load_rates():
    with open(RATES_JSON) as f:
        return json.load(f)


def find_rate(rates, service_substr, meter_substr):
    """Return the FIRST rates.json index whose service+meter both contain the given
    case-insensitive substrings. Raises if not found -- fail loud on a broken lookup."""
    s, m = service_substr.lower(), meter_substr.lower()
    for i, r in enumerate(rates):
        if s in r["service"].lower() and m in r["meter"].lower():
            return i
    raise ValueError(
        f"No rate matched service={service_substr!r} meter={meter_substr!r}"
    )


def resolve_named_rates(rates, rate_cells):
    """One shared lookup table (friendly name -> 'Rates!$C$n') consumed by both the
    Pod Build and Monthly TCO sheets, so every cost formula points at the same cell."""

    def rc(service_substr, meter_substr):
        return rate_cells[find_rate(rates, service_substr, meter_substr)]

    return {
        "api_gw": rc("API Gateway", "ApiGatewayRequest"),
        "lambda_gbs": rc("Lambda", "GB-Second"),
        "lambda_req": rc("Lambda", "Request"),
        "cw_ingest": rc("CloudWatch", "DataProcessing"),
        "cw_storage": rc("CloudWatch", "TimedStorage"),
        "cw_metric": rc("CloudWatch", "MetricMonitorUsage"),
        "cw_dashboard": rc("CloudWatch", "Dashboard"),
        "s3_storage": rc("S3 Standard", "TimedStorage"),
        "s3_put": rc("S3 Standard", "Tier1"),
        "s3_get": rc("S3 Standard", "Tier2"),
        "eb_scheduler": rc("EventBridge Scheduler", "ScheduledInvocation"),
        "eventbridge": rc("EventBridge", "64K"),
        "sqs": rc("SQS", "Requests-RBP"),
        "opensearch_instance": rc("OpenSearch Service", "r7g.large"),
        "opensearch_storage": rc("OpenSearch Service (managed", "gp3"),
        "titan_in": rc("Titan", "on-demand"),
        "fargate_vcpu": rc("Fargate (x86", "vCPU-hour"),
        "fargate_gb": rc("Fargate (x86", "GB-hour"),
        "aurora_2xlarge": rc(
            "Aurora PostgreSQL", "db.r6g.2xlarge Single-AZ instance (Standard)"
        ),
        "aurora_xlarge": rc(
            "Aurora PostgreSQL", "db.r6g.xlarge Single-AZ instance (Standard)"
        ),
        "aurora_large": rc(
            "Aurora PostgreSQL", "db.r6g.large Single-AZ instance (Standard)"
        ),
        "agentcore_st": rc("AgentCore Memory", "Short-term"),
        "agentcore_lt_store": rc("AgentCore Memory", "built-in strategies"),
        "agentcore_lt_retrieve": rc("AgentCore Memory", "retrieval"),
        "opus_in": rc("Opus 4.8", "Input tokens (cross"),
        "opus_out": rc("Opus 4.8", "Output tokens"),
        "sonnet_in": rc("Sonnet 5", "Input tokens (cross"),
        "sonnet_out": rc("Sonnet 5", "Output tokens"),
        "haiku_in": rc("Haiku 4.5", "Input tokens (cross"),
        "haiku_out": rc("Haiku 4.5", "Output tokens"),
        "appsync_ops": rc("AppSync", "GraphQLInvocation"),
        "cloudfront_req": rc("CloudFront", "Requests-Tier2"),
        "cloudfront_out": rc("CloudFront", "DataTransfer-Out"),
        "alb_hour": rc("Load Balancer", "LoadBalancerUsage"),
        "alb_lcu": rc("Load Balancer", "LCUUsage"),
        "nat_hour": rc("NAT Gateway", "Hourly"),
        "nat_gb": rc("NAT Gateway", "per GB"),
        "secrets": rc("Secrets Manager", "Secrets"),
        "vpc_hour": rc("VPC Interface", "Hours"),
        "vpc_gb": rc("VPC Interface", "Bytes"),
    }


def sku_gate(model_sel, opus_expr, sonnet_expr, haiku_expr):
    """SEGRO-style SKU zero-gate: nested IF on ModelSKU so only the selected model prices."""
    return (
        f'IF({model_sel}="Opus-4.8",{opus_expr},'
        f'IF({model_sel}="Sonnet-5",{sonnet_expr},{haiku_expr}))'
    )


CP_SHEET = "Control Panel >>>"

# (row, name, default, kind, validation, drives_note)
CP_INPUTS = [
    (
        5,
        "DeliveriesPerMonth",
        10,
        "qty",
        ("decimal", ">=", 0),
        "Vendor deliveries per month -> volume model",
    ),
    (
        6,
        "ProductsPerDelivery",
        989,
        "qty",
        ("decimal", ">=", 1),
        "Products per delivery -> ItemsPerMonth = Deliveries x Products",
    ),
    (
        7,
        "DatasetCount",
        5000,
        "qty",
        ("decimal", ">=", 1),
        "Catalogue size (CDAO datasets) -> storage lines",
    ),
    (
        8,
        "LlmReachablePct",
        1.0,
        "pct",
        ("decimal", "between", (0, 1)),
        "Fraction of items reaching the agent loop. CODE-VERIFIED default: the deployed "
        "Strands orchestrator invokes specialist+verifier per product UNCONDITIONALLY "
        "(not band-gated) -> 1.0, DERIVED. Lower it to model a band-gated future.",
    ),
    (
        9,
        "ModelSKU",
        "Opus-4.8",
        "list",
        ("list", None, "Opus-4.8,Sonnet-5,Haiku-4.5"),
        "Bedrock model SKU selection via SKU-gate pattern (vendor fixed = Bedrock)",
    ),
    (
        10,
        "InputTokensPerCallK",
        0.8,
        "tok",
        ("decimal", ">=", 0),
        "Per-call input tokens (thousands). CODE-VERIFIED LIKELY ~0.69-0.94K (specialist), "
        "~0.38-0.52K (verifier). Client sketch said 50K -> carried as the HIGH scenario row, "
        "not the default.",
    ),
    (
        11,
        "OutputTokensPerCallK",
        0.25,
        "tok",
        ("decimal", ">=", 0),
        "Per-call output tokens (thousands). CODE-VERIFIED LIKELY ~0.21-0.27K.",
    ),
    (
        12,
        "CallsPerItem",
        2,
        "qty",
        ("decimal", ">=", 0),
        "Bedrock calls per item. CODE-VERIFIED: 2 (mapping specialist + verifier). Dense-arm "
        "opus_dense_score is OFF by default (DenseBackend=jaro_winkler); if flipped, it fires "
        "per taxonomy node scored -> carried as a separate optional line, default 0.",
    ),
    (
        13,
        "PointRate",
        0,
        "money2",
        ("decimal", ">=", 0),
        "$ per build point. DEFAULT 0 = RATE_MISSING visible zero (un-costed, not free).",
    ),
    (
        14,
        "EnvMultiplier",
        1,
        "list",
        ("list", None, "1,2,3"),
        "Non-prod environment multiplier on always-on infra",
    ),
    (
        15,
        "AuroraSize",
        "2xlarge",
        "list",
        ("list", None, "large,xlarge,2xlarge"),
        "Aurora instance class selector (client sketch says 2xLarge; repo-real default is "
        "Serverless v2 db.serverless, a material sizing divergence -- see Assumptions)",
    ),
    (
        16,
        "HitlPct",
        0.15,
        "pct",
        ("decimal", "between", (0, 1)),
        "Fraction of items routed to HITL review (AppSync traffic)",
    ),
    (
        17,
        "StoragePerItemKB",
        50,
        "qty",
        ("decimal", ">=", 0),
        "Canonical JSON-LD per product (repo estimate: ~2-4KB actual; 50KB is a conservative "
        "client-facing ceiling -- see Assumptions)",
    ),
    (
        18,
        "SlipWeeks",
        0,
        "qty",
        ("decimal", ">=", 0),
        "Gantt stretch (14-week core, spec 5b with 29->14)",
    ),
    (
        19,
        "ContingencyPct",
        0.15,
        "pct",
        ("decimal", ">=", 0),
        "Uplift on monthly TCO subtotal",
    ),
]

FMT_BY_KIND = {
    "qty": QTY_FMT,
    "pct": PCT_FMT,
    "tok": "0.000",
    "money2": MONEY2_FMT,
    "list": "General",
}


def build_control_panel(wb):
    ws = wb.create_sheet(CP_SHEET)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 84

    style_title(
        ws, "D", "JPMC SCUDO POAP TCO -- Control Panel (all scenarios drive live)"
    )
    ws["A2"] = (
        "Amber cells are editable inputs -- change a value and every tab, total and "
        "chart recalculates live. Status labels: VERIFIED (grounded to source) / "
        "INDICATIVE (list price, editable) / DERIVED (formula) / ASSUMPTION (stated "
        "basis, editable) / RATE_MISSING (no source rate -- $0 until entered). USD, "
        "us-east-1 throughout."
    )
    ws.merge_cells("A2:D2")
    ws["A2"].font = Font(italic=True, color="FF808080", size=11)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 30

    style_section(ws, 4, "D", "SCENARIO SWITCHES & LOADINGS")

    for row, name, default, kind, validation, note in CP_INPUTS:
        ws[f"A{row}"] = name
        cell = ws[f"B{row}"]
        cell.value = default
        cell.fill = AMBER_FILL
        cell.font = BLUE
        if kind != "list":
            cell.number_format = FMT_BY_KIND[kind]
        status_cell(
            ws, f"C{row}", "ASSUMPTION" if name != "LlmReachablePct" else "DERIVED"
        )
        ws[f"D{row}"] = note
        ws[f"D{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        vtype, op, arg = validation
        dv = None
        if vtype == "decimal" and op == ">=":
            dv = DataValidation(
                type="decimal", operator="greaterThanOrEqual", formula1=str(arg)
            )
        elif vtype == "decimal" and op == "between":
            dv = DataValidation(
                type="decimal",
                operator="between",
                formula1=str(arg[0]),
                formula2=str(arg[1]),
            )
        elif vtype == "list":
            dv = DataValidation(type="list", formula1=f'"{arg}"')
        if dv is not None:
            dv.add(cell)
            ws.add_data_validation(dv)

        wb_ = wb
        wb_.defined_names[name] = DefinedName(name, attr_text=f"'{CP_SHEET}'!$B${row}")

    style_section(ws, 21, "D", "DERIVED VOLUME")
    ws["A22"] = "Items per month (Deliveries x Products)"
    ws["B22"] = "=DeliveriesPerMonth*ProductsPerDelivery"
    ws["B22"].number_format = QTY_FMT
    status_cell(ws, "C22", "DERIVED")
    ws["A23"] = "LLM-reachable items per month"
    ws["B23"] = "=ItemsPerMonth*LlmReachablePct"
    ws["B23"].number_format = QTY_FMT
    status_cell(ws, "C23", "DERIVED")
    wb.defined_names["ItemsPerMonth"] = DefinedName(
        "ItemsPerMonth", attr_text=f"'{CP_SHEET}'!$B$22"
    )
    wb.defined_names["LlmItems"] = DefinedName(
        "LlmItems", attr_text=f"'{CP_SHEET}'!$B$23"
    )

    notes = [
        "LlmReachablePct default 1.0 is CODE-VERIFIED (orchestrator.py:113-151): the deployed "
        "Strands orchestrator has no dependency on the confidence-band gate for LLM reachability "
        "-- it calls mapping+verifier for every non-RESEARCH product. Lower this to model a "
        "future band-gated design.",
        "InputTokensPerCallK/OutputTokensPerCallK defaults are the CODE-VERIFIED LIKELY scenario "
        "(measured prompt-template character counts, llm_call_model.md SS3). The client sketch's "
        "50K in / 1K out is carried as a labelled CLIENT-SKETCH HIGH scenario row on Monthly TCO, "
        "not the default.",
        "PointRate defaults to 0 by design (RATE_MISSING convention): Pod Build shows visible $0 "
        "build cost until a real $/point rate is supplied -- this is un-costed, not free.",
        "AuroraSize offers large/xlarge/2xlarge fixed-instance rates from Rates!, matching the "
        "client sketch's 'Aurora db 2xLarge' line. The repo-real deployed config is Aurora "
        "Serverless v2 (db.serverless, 0.5-2 ACU) -- a material sizing divergence, see Assumptions.",
        "StoragePerItemKB defaults to 50KB as a conservative client-facing ceiling. The repo-real "
        "measured canonical JSON-LD payload is ~2-4KB per fully-published product "
        "(arch_infra_persistence.md SSC.5) -- storage cost at 50KB is a deliberate overstatement, "
        "not a repo-measured figure.",
        "All figures are USD, AWS us-east-1 region only -- no other currency or cloud vendor.",
    ]
    r = 25
    ws[f"A{r}"] = "SCOPE NOTES"
    ws[f"A{r}"].font = WHITE_BOLD
    ws[f"A{r}"].fill = SECTION_FILL
    ws.merge_cells(f"A{r}:D{r}")
    for i, note in enumerate(notes, start=1):
        row = r + i
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = f"* {note}"
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 28

    ws.freeze_panes = "A3"


def _value_fmt(usd):
    if usd == 0:
        return MONEY_FMT
    if usd < 0.001:
        return MONEY8_FMT
    if usd < 1:
        return MONEY4_FMT
    return MONEY2_FMT


def build_rates(wb, rates):
    """Write every rates.json meter, grouped by service. Returns {index: 'Rates!$C$<row>'}."""
    ws = wb.create_sheet("Rates")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 62
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 60

    style_title(
        ws, "F", "Rates -- USD list prices, us-east-1, on-demand (July 2026)", height=38
    )
    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "Every figure on this tab is an editable input. All rates VERIFIED from AWS Price "
        "List Bulk API / aws.amazon.com/*/pricing pages, us-east-1, as cited per row. "
        "RATE_MISSING lines are $0 until a real rate is supplied. USD only."
    )
    ws["A2"].font = Font(italic=True, color="FF808080", size=11)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 30

    headers = ["Rate", "Unit", "Value (USD)", "Status", "Source URL", "Notes / As-of"]
    row = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    cellmap = {}
    last_service = None
    for i, r in enumerate(rates):
        if r["service"] != last_service:
            ws.merge_cells(f"A{row}:F{row}")
            ws[f"A{row}"] = r["service"].upper()
            ws[f"A{row}"].font = WHITE_BOLD
            ws[f"A{row}"].fill = SECTION_FILL
            last_service = r["service"]
            row += 1
        ws[f"A{row}"] = r["meter"]
        ws[f"B{row}"] = r["unit"]
        val_cell = ws[f"C{row}"]
        val_cell.value = r["usd"]
        val_cell.font = BLUE
        val_cell.number_format = _value_fmt(r["usd"])
        status_cell(ws, f"D{row}", r["confidence"])
        ws[f"E{row}"] = r["source_url"]
        ws[f"E{row}"].font = GREY
        ws[f"F{row}"] = r["notes"]
        ws[f"F{row}"].font = GREY
        ws[f"F{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        cellmap[i] = f"Rates!$C${row}"
        row += 1

    ws.freeze_panes = "A5"
    return cellmap


POD_COLS = [
    "Pod",
    "Short Title",
    "AWS Components",
    "Points",
    "Run freq per mo",
    "Est data vol GB",
    "Input tokens K",
    "Output tokens K",
    "Cost estimate $ monthly",
    "Build cost (Points x PointRate)",
    "Status",
    "Basis / source",
]

# Each row: (pod, short_title, aws_components, points, cost_kind, cost_args, status, basis)
# cost_kind drives the Excel-formula generator in build_pod_build().
POD_ROWS_A = [
    (
        "A",
        "API GW entry + routing",
        "API Gateway (HttpApi)",
        2,
        "api_gw",
        {},
        "ASSUMPTION",
        "Points: relative sizing, no client estimate supplied. Cost: ItemsPerMonth requests "
        "through the /run route, ApiGatewayRequest rate. template.yaml:589-599.",
    ),
    (
        "A",
        "Lambda authorizer",
        "Lambda (auth check)",
        1,
        "zero",
        {},
        "ASSUMPTION",
        "No JWT Lambda authorizer exists in the repo today (auth is an API-key header) -- "
        "arch_infra_persistence.md D.1. RATE_MISSING: $0, un-costed, not free, until this "
        "component is actually built.",
    ),
    (
        "A",
        "Ingestion Coordinator (ETL worker)",
        "Lambda EtlFn, 3008MB, SQS-triggered",
        3,
        "lambda",
        {"mem_gb": 3008 / 1024, "dur_s": 2.0},
        "DERIVED",
        "template.yaml:16-25 EtlFn 3008MB default memory, SQS BatchSize 5. Duration ASSUMPTION "
        "(~2s/product, no repo timing evidence); GB-second + per-request Lambda rates.",
    ),
    (
        "A",
        "Sanity Check (validate/quarantine)",
        "In-process in EtlFn -- no separate Lambda",
        1,
        "zero",
        {},
        "DERIVED",
        "etl_handler.py _process_object/_quarantine run inside EtlFn -- no incremental "
        "compute cost beyond the Ingestion Coordinator line above.",
    ),
    (
        "A",
        "CloudWatch (logs + metrics)",
        "CloudWatch log groups, EMF metrics, dashboard",
        1,
        "cloudwatch",
        {"log_gb_per_item_kb": 2, "n_metrics": 15, "n_dashboards": 1},
        "ASSUMPTION",
        "~2KB structured log per product ASSUMPTION; 15 custom EMF metrics per MEMORY note; "
        "1 dashboard. arch_infra_persistence.md A.8.",
    ),
    (
        "A",
        "S3 raw landing",
        "S3 RawFeedBucket",
        1,
        "s3_put_get",
        {"puts_per_item": 1, "gets_per_item": 0},
        "DERIVED",
        "template.yaml:176-198 RawFeedBucket, EventBridge-notification-enabled. 1 PUT/vendor "
        "object at ingest, storage sized via StoragePerItemKB.",
    ),
    (
        "A",
        "EventBridge Scheduler (poller)",
        "EventBridge Scheduler, cron twice-monthly",
        1,
        "eb_scheduler",
        {"invocations_per_month": 2},
        "VERIFIED",
        "PollerFn cron(0 6 1,15 * ? *) -- twice-monthly, not per-product. "
        "arch_infra_persistence.md A.1/A.3.",
    ),
    (
        "A",
        "EventBridge ObjectCreated rule",
        "EventBridge rule, S3->SQS",
        1,
        "eventbridge",
        {"events_per_item": 1},
        "VERIFIED",
        "RawFeedIngressRule on default bus, S3 Object Created -> EtlQueue. "
        "arch_infra_persistence.md A.3.",
    ),
    (
        "A",
        "SQS (EtlQueue + DLQ)",
        "SQS standard queue + DLQ",
        1,
        "sqs",
        {"messages_per_item": 1},
        "VERIFIED",
        "EtlQueue BatchSize 5, DLQ after 3 receives. template.yaml A.2.",
    ),
    (
        "A",
        "ETL worker VPC egress (NAT)",
        "NAT Gateway (no-VPC EtlFn needs internet egress)",
        1,
        "nat_share",
        {},
        "DERIVED",
        "EtlFn explicitly has NO VPC (template.yaml:488-491); NAT cost is carried once as a "
        "shared always-on line under Pod E Platform, not duplicated here -- see Pod E.",
    ),
    (
        "A",
        "S3 canonical sink",
        "S3 CleanCanonicalBucket",
        1,
        "s3_put_get",
        {"puts_per_item": 1, "gets_per_item": 0},
        "DERIVED",
        "CleanCanonicalBucket, written by etl_handler.py pass path. 1 PUT/product.",
    ),
    (
        "A",
        "S3 quarantine",
        "S3 QuarantineBucket",
        1,
        "s3_put_get",
        {"puts_per_item": 0, "gets_per_item": 0},
        "ASSUMPTION",
        "QuarantineBucket, mutually exclusive with canonical PUT (bad-file path only) -- "
        "modelled at $0 incremental (rare-path, no repo-measured quarantine rate).",
    ),
]

POD_ROWS_B = [
    (
        "B",
        "AIA Matching Engine gate A (scope)",
        "In-process, matching.py",
        2,
        "zero",
        {},
        "DERIVED",
        "Deterministic Python gate, no separate compute -- folded into the matching Lambda "
        "compute already priced on the Strands Orchestrator line (Pod C). ZONES.md Zone 3.",
    ),
    (
        "B",
        "AIA Matching Engine gate B (precedent)",
        "In-process, aurora_memory.py CONSULT",
        2,
        "aurora_ops",
        {"ops_per_item": 1},
        "RATE_MISSING",
        "1 Aurora RDS Data API ExecuteStatement per item for precedent lookup "
        "(arch_infra_persistence.md C.4 roll-up), but rates.json has no per-request RDS "
        "Data API meter -- the underlying Aurora instance-hour cost is already priced on "
        "the Aurora 2xlarge line below; this row's incremental per-call charge is $0, "
        "un-costed not free, until a Data API request rate is sourced.",
    ),
    (
        "B",
        "AIA Matching Engine gate C (hybrid retrieval)",
        "FalkorDB on ECS (retrieval store)",
        3,
        "falkordb_share",
        {},
        "DERIVED",
        "FalkorDB on Fargate is the live retrieval store (both deploy paths) -- always-on ECS "
        "task cost carried once under Pod E Platform, not duplicated per-pod. "
        "arch_infra_persistence.md A.8.",
    ),
    (
        "B",
        "Titan Embeddings (indexing/write path)",
        "Bedrock Titan Text Embeddings V2",
        2,
        "titan",
        {"tokens_per_item_k": 0.5},
        "VERIFIED",
        "Titan IS called from projection_handler.py:483 on the write/index path only (not "
        "retrieval, which is parked). ~0.5K tokens/item ASSUMPTION for a short product "
        "description embed.",
    ),
    (
        "B",
        "OpenSearch fuzzy+kNN",
        "OpenSearch Service r7g.large.search + gp3 storage",
        4,
        "opensearch",
        {},
        "DERIVED",
        "Template-real domain; write path real (_index_opensearch), matching read path today "
        "is a Jaro-Winkler/FalkorDB seam not OpenSearch kNN. Priced at the always-on domain "
        "cost the client's list implies. arch_infra_persistence.md A.7/A.8.",
    ),
    (
        "B",
        "Aurora 2xlarge (system of record)",
        "Aurora PostgreSQL, AuroraSize selector",
        5,
        "aurora_instance",
        {},
        "DERIVED",
        "Client sketch says 'Aurora db 2xLarge'; repo-real deployed shape is Aurora Serverless "
        "v2 (db.serverless) -- material divergence, see Assumptions. This line prices the "
        "client's fixed-instance-class ask via the AuroraSize selector.",
    ),
]

POD_ROWS_C = [
    (
        "C",
        "Strands Orchestrator Opus 4.8",
        "Lambda ScudoFn (3008MB) + Bedrock routing",
        5,
        "orchestrator_lambda",
        {"mem_gb": 3008 / 1024, "dur_s": 1.0},
        "VERIFIED",
        "ScudoFn 3008MB default, no LLM call in Orchestrator.route() itself (deterministic "
        "Python). orchestrator.py:102-110; llm_call_model.md SS1.",
    ),
    (
        "C",
        "AgentCore Memory (short-term events)",
        "Bedrock AgentCore Memory meter (bespoke store)",
        3,
        "agentcore_st",
        {"events_per_item": 2},
        "VERIFIED",
        "Client wants this priced; repo-real substrate is aurora_memory.py (bespoke), not the "
        "managed AgentCore Memory service (arch_infra_persistence.md D.1). Rate from Bedrock "
        "AgentCore Memory short-term-events meter, applied notionally as the client-facing "
        "unit-cost analog.",
    ),
    (
        "C",
        "AgentCore Memory (long-term records)",
        "Bedrock AgentCore Memory, built-in strategies",
        2,
        "agentcore_lt_store",
        {"records_per_item": 1},
        "ASSUMPTION",
        "1 long-term record/item ASSUMPTION (no repo-measured record count); rate VERIFIED "
        "from rates.json.",
    ),
    (
        "C",
        "AgentCore Memory (long-term retrieval)",
        "Bedrock AgentCore Memory retrieval meter",
        2,
        "agentcore_lt_retrieve",
        {"retrievals_per_item": 1},
        "ASSUMPTION",
        "1 retrieval/item ASSUMPTION; rate VERIFIED from rates.json.",
    ),
    (
        "C",
        "Specialist agent (mapping)",
        "Bedrock InvokeModel, ModelSKU-gated",
        4,
        "llm_specialist",
        {},
        "VERIFIED",
        "1 InvokeModel call/product via _call_mapping (orchestrator.py:191-203). SKU-gated "
        "token cost -- see Monthly TCO for the full CODE-VERIFIED LIKELY / CLIENT-SKETCH HIGH "
        "scenario rows.",
    ),
    (
        "C",
        "Verifier agent 10-dim",
        "Bedrock InvokeModel, ModelSKU-gated",
        4,
        "llm_verifier",
        {},
        "VERIFIED",
        "1 InvokeModel call/product via _call_verifier (orchestrator.py:237-261), 10-dim "
        "rubric, VerifierReport min_length=10/max_length=10. schemas.py:219-247.",
    ),
    (
        "C",
        "Dense-arm optional (opus_dense per node)",
        "Bedrock InvokeModel, opt-in, default OFF",
        3,
        "llm_dense_optional",
        {},
        "DERIVED",
        "Only fires if SCUDO_DENSE_BACKEND=opus (default jaro_winkler, infra/scudo-poc-app.yaml"
        ":76-79) -- OFF by default, amber multiplier default 0. When on: N calls/product, "
        "N=taxonomy nodes scored (14 in the demo fixture). llm_call_model.md SS1/SS3.",
    ),
    (
        "C",
        "Bedrock Evaluations (judge)",
        "LLM-as-judge, billed as judge model's own tokens",
        2,
        "zero",
        {},
        "ASSUMPTION",
        "No separate meter for LLM-as-judge (rates.json: billed as judge model's standard "
        "tokens). Repo has no real Bedrock Evaluations API call anywhere -- bespoke offline "
        "harness instead (arch_infra_persistence.md D.1). $0 incremental line, human-eval "
        "meter available in Rates if wired later.",
    ),
]

POD_ROWS_D = [
    (
        "D",
        "JAPI Persist outbox writer",
        "Aurora ExecuteStatement (RDS Data API)",
        5,
        "aurora_ops",
        {"calls_per_item": 10},
        "RATE_MISSING",
        "~10 ExecuteStatement calls/fully-published product (outbox insert + projection "
        "rows) -- VERIFIED call count, arch_infra_persistence.md SSC, persistence-ops-"
        "per-product roll-up. No per-request RDS Data API rate in rates.json, so the "
        "incremental per-call charge is $0 (un-costed not free) until sourced; the Aurora "
        "instance-hour cost itself is priced separately on the Aurora 2xlarge line.",
    ),
    (
        "D",
        "Outbox sweep (EventBridge Schedule)",
        "rate(5 min) Scheduler -> Lambda sweep",
        3,
        "eb_scheduler",
        {"invokes_per_month": 8928},
        "VERIFIED",
        "rate(5 min) = 12/hr * 24 * mean(30.4 days) ~= 8928 invokes/mo, flat regardless of "
        "volume. projection_handler.py sweep_outbox; infra/HANDOVER_5zone_alignment.md.",
    ),
    (
        "D",
        "HITL AppSync API",
        "AppSync GraphQL ops + notifications",
        4,
        "appsync",
        {"ops_per_hitl_item": 6},
        "ASSUMPTION",
        "6 GraphQL ops/HITL item (list+get+mutate+subscribe-notify) is an ASSUMPTION -- no "
        "repo-measured op count; HitlPct selector gates volume. rate VERIFIED from rates.json.",
    ),
    (
        "D",
        "HITL review ECS task (reviewer console)",
        "Fargate, always-on shared task",
        4,
        "ecs_share",
        {"vcpu": 0.5, "mem_gb": 1.0},
        "DERIVED",
        "Shared with Pod E platform ECS allocation -- this row shows the HITL-attributable "
        "slice only, full task priced once under Pod E to avoid double count.",
    ),
    (
        "D",
        "S3 projection sink + CloudFront invalidation",
        "S3 PUT/GET + CloudFront requests",
        2,
        "s3_put_get",
        {"puts_per_item": 1, "gets_per_item": 2},
        "ASSUMPTION",
        "1 PUT (projection write) + 2 GET (dashboard reads)/item ASSUMPTION.",
    ),
]

POD_ROWS_E = [
    (
        "E",
        "FalkorDB on ECS (retrieval store)",
        "Fargate task, always-on, shared",
        5,
        "falkordb_share",
        {"vcpu": 1.0, "mem_gb": 2.0},
        "VERIFIED",
        "Actual live hybrid-retrieval substrate (Neptune/OpenSearch-kNN are dormant seams). "
        "1 vCPU / 2GB always-on Fargate task. ZONES.md Zone3; ua-dogfood-scudo memory.",
    ),
    (
        "E",
        "CloudFront distribution",
        "Requests + data transfer out, PoC demo",
        4,
        "cloudfront",
        {"requests_per_month": 50000, "gb_out": 10},
        "ASSUMPTION",
        "50k requests / 10GB out per month is an ASSUMPTION sized for a PoC-scale demo, not "
        "production traffic -- see Assumptions sheet.",
    ),
    (
        "E",
        "ALB (console + API ingress)",
        "Hourly + LCU-hours, always-on",
        3,
        "alb",
        {},
        "VERIFIED",
        "1 ALB, always-on hourly charge + nominal LCU-hours. infra/scudo-poc-*.yaml.",
    ),
    (
        "E",
        "NAT Gateway (VPC egress)",
        "Hourly + per-GB processed, always-on",
        3,
        "nat_gateway",
        {"gb_per_month": 20},
        "ASSUMPTION",
        "20GB/mo egress ASSUMPTION (ETL worker + Lambda-in-VPC egress); this is the priced "
        "line that Pod A's nat_share and Pod B's aurora_ops VPC path reference.",
    ),
    (
        "E",
        "Secrets Manager (DB + API creds)",
        "Per-secret + API-call meter",
        1,
        "secrets_manager",
        {"secret_count": 4},
        "VERIFIED",
        "4 secrets (CONSOLE_DB_PASSWORD, Bedrock/API keys, MFT gateway cred, JWT signing key).",
    ),
    (
        "E",
        "VPC Interface Endpoints",
        "Per-endpoint-hour + per-GB, always-on",
        2,
        "vpc_endpoints",
        {"endpoint_count": 3},
        "ASSUMPTION",
        "3 endpoints ASSUMPTION (Secrets Manager, Bedrock, S3-gateway-equivalent) -- no "
        "repo-measured endpoint count.",
    ),
    (
        "E",
        "CodeBuild (console/dashboard build)",
        "Build-minute meter, deploy-time only",
        1,
        "zero",
        {},
        "RATE_MISSING",
        "No CodeBuild meter in Rates -- research/rates.json has zero CodeBuild entries. "
        "$0 visible, un-costed not free, until a real build-minute rate is sourced. "
        "scudo-poc-console-build CodeBuild project exists in-repo (deploy-time only, "
        "not steady-state run cost).",
    ),
]


def _pod_cost_formula(kind, params, nr):
    """Return (freq_txt, data_vol_txt, tok_in_txt, tok_out_txt, cost_formula) for one
    Pod Build row. cost_formula always starts with '=' and references named ranges /
    Rates! cells -- never a bare Python-computed constant."""
    NA = "-"
    if kind == "zero" or kind == "nat_share" or kind == "aurora_ops":
        return ("-", NA, NA, NA, "=ItemsPerMonth*0")
    if kind == "falkordb_share" and "vcpu" not in params:
        return ("Always-on (shared, see Pod E)", NA, NA, NA, "=ItemsPerMonth*0")
    if kind == "api_gw":
        return (
            "=ItemsPerMonth",
            NA,
            NA,
            NA,
            f"=ItemsPerMonth*{nr['api_gw']}/1000000",
        )
    if kind in ("lambda", "orchestrator_lambda"):
        mem_gb, dur_s = params["mem_gb"], params["dur_s"]
        return (
            "=ItemsPerMonth",
            NA,
            NA,
            NA,
            f"=ItemsPerMonth*{mem_gb}*{dur_s}*{nr['lambda_gbs']}"
            f"+ItemsPerMonth*{nr['lambda_req']}/1000000",
        )
    if kind == "cloudwatch":
        log_kb, n_metrics, n_dash = (
            params["log_gb_per_item_kb"],
            params["n_metrics"],
            params["n_dashboards"],
        )
        return (
            "=ItemsPerMonth",
            "=ItemsPerMonth*{}/1000000".format(log_kb),
            NA,
            NA,
            f"=(ItemsPerMonth*{log_kb}/1000000)*({nr['cw_ingest']}+{nr['cw_storage']})"
            f"+{n_metrics}*{nr['cw_metric']}+{n_dash}*{nr['cw_dashboard']}",
        )
    if kind == "s3_put_get":
        puts, gets = params["puts_per_item"], params["gets_per_item"]
        storage = (
            f"+ItemsPerMonth*StoragePerItemKB/1000000*{nr['s3_storage']}"
            if puts > 0
            else ""
        )
        data_vol = "=ItemsPerMonth*StoragePerItemKB/1000000" if puts > 0 else NA
        return (
            "=ItemsPerMonth",
            data_vol,
            NA,
            NA,
            f"=ItemsPerMonth*{puts}*{nr['s3_put']}/1000"
            f"+ItemsPerMonth*{gets}*{nr['s3_get']}/1000{storage}",
        )
    if kind == "eb_scheduler":
        n = params.get("invocations_per_month", params.get("invokes_per_month"))
        return (
            str(n),
            NA,
            NA,
            NA,
            f"={n}*{nr['eb_scheduler']}/1000000",
        )
    if kind == "eventbridge":
        n = params["events_per_item"]
        return (
            "=ItemsPerMonth",
            NA,
            NA,
            NA,
            f"=ItemsPerMonth*{n}*{nr['eventbridge']}/1000000",
        )
    if kind == "sqs":
        n = params["messages_per_item"]
        return (
            "=ItemsPerMonth",
            NA,
            NA,
            NA,
            f"=ItemsPerMonth*{n}*{nr['sqs']}/1000000",
        )
    if kind == "aurora_instance":
        return (
            "Always-on",
            NA,
            NA,
            NA,
            f'=730*EnvMultiplier*IF(AuroraSize="large",{nr["aurora_large"]},'
            f'IF(AuroraSize="xlarge",{nr["aurora_xlarge"]},{nr["aurora_2xlarge"]}))',
        )
    if kind == "titan":
        tok_k = params["tokens_per_item_k"]
        return (
            "=LlmItems",
            NA,
            str(tok_k),
            "0",
            f"=LlmItems*{tok_k}*{nr['titan_in']}/1000",
        )
    if kind == "opensearch":
        return (
            "Always-on",
            "=DatasetCount*StoragePerItemKB/1000000",
            NA,
            NA,
            f"=730*EnvMultiplier*{nr['opensearch_instance']}"
            f"+DatasetCount*StoragePerItemKB/1000000*{nr['opensearch_storage']}",
        )
    if kind == "falkordb_share":  # Pod E actual line (params has vcpu/mem_gb)
        vcpu, mem_gb = params["vcpu"], params["mem_gb"]
        return (
            "Always-on",
            NA,
            NA,
            NA,
            f"=730*EnvMultiplier*({vcpu}*{nr['fargate_vcpu']}+{mem_gb}*{nr['fargate_gb']})",
        )
    if kind in ("llm_specialist", "llm_verifier"):
        sku_in = sku_gate("ModelSKU", nr["opus_in"], nr["sonnet_in"], nr["haiku_in"])
        sku_out = sku_gate(
            "ModelSKU", nr["opus_out"], nr["sonnet_out"], nr["haiku_out"]
        )
        return (
            "=LlmItems",
            NA,
            "=InputTokensPerCallK",
            "=OutputTokensPerCallK",
            f"=LlmItems*1*(InputTokensPerCallK/1000*{sku_in}"
            f"+OutputTokensPerCallK/1000*{sku_out})",
        )
    if kind == "llm_dense_optional":
        # amber optional multiplier lives on this row itself (default 0, off) -- see
        # caller, which writes the amber cell and passes its ref in params["_cell"].
        cell = params["_cell"]
        sku_in = sku_gate("ModelSKU", nr["opus_in"], nr["sonnet_in"], nr["haiku_in"])
        sku_out = sku_gate(
            "ModelSKU", nr["opus_out"], nr["sonnet_out"], nr["haiku_out"]
        )
        return (
            f"=ItemsPerMonth*{cell}",
            NA,
            "0.5",
            "0.08",
            f"=ItemsPerMonth*{cell}*(0.5/1000*{sku_in}+0.08/1000*{sku_out})",
        )
    if kind == "agentcore_st":
        n = params["events_per_item"]
        return (
            "=LlmItems",
            NA,
            NA,
            NA,
            f"=LlmItems*{n}*{nr['agentcore_st']}/1000",
        )
    if kind == "agentcore_lt_store":
        n = params["records_per_item"]
        return (
            "=LlmItems",
            NA,
            NA,
            NA,
            f"=LlmItems*{n}*{nr['agentcore_lt_store']}/1000",
        )
    if kind == "agentcore_lt_retrieve":
        n = params["retrievals_per_item"]
        return (
            "=LlmItems",
            NA,
            NA,
            NA,
            f"=LlmItems*{n}*{nr['agentcore_lt_retrieve']}/1000",
        )
    if kind == "appsync":
        n = params["ops_per_hitl_item"]
        return (
            "=ItemsPerMonth*HitlPct",
            NA,
            NA,
            NA,
            f"=ItemsPerMonth*HitlPct*{n}*{nr['appsync_ops']}/1000000",
        )
    if kind == "ecs_share":
        vcpu, mem_gb = params["vcpu"], params["mem_gb"]
        return (
            "Always-on",
            NA,
            NA,
            NA,
            f"=730*EnvMultiplier*({vcpu}*{nr['fargate_vcpu']}+{mem_gb}*{nr['fargate_gb']})",
        )
    if kind == "cloudfront":
        req, gb = params["requests_per_month"], params["gb_out"]
        return (
            str(req),
            str(gb),
            NA,
            NA,
            f"={req}*{nr['cloudfront_req']}/10000+{gb}*{nr['cloudfront_out']}",
        )
    if kind == "alb":
        return (
            "Always-on",
            NA,
            NA,
            NA,
            f"=730*EnvMultiplier*{nr['alb_hour']}+730*EnvMultiplier*1*{nr['alb_lcu']}",
        )
    if kind == "nat_gateway":
        gb = params["gb_per_month"]
        return (
            "Always-on",
            str(gb),
            NA,
            NA,
            f"=730*EnvMultiplier*{nr['nat_hour']}+{gb}*{nr['nat_gb']}",
        )
    if kind == "secrets_manager":
        n = params["secret_count"]
        return ("Always-on", NA, NA, NA, f"={n}*{nr['secrets']}")
    if kind == "vpc_endpoints":
        n = params["endpoint_count"]
        return (
            "Always-on",
            NA,
            NA,
            NA,
            f"={n}*730*EnvMultiplier*{nr['vpc_hour']}",
        )
    raise ValueError(f"Unhandled Pod Build cost_kind: {kind!r}")


def build_pod_build(wb, rates, rate_cells, assumptions_log):
    nr = resolve_named_rates(rates, rate_cells)
    ws = wb.create_sheet("Pod Build")
    ws.sheet_view.showGridLines = False
    widths = [6, 34, 40, 8, 20, 16, 14, 14, 20, 22, 12, 60, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    style_title(
        ws, "L", "Pod Build -- client table shape + repo-real extras", height=18
    )
    style_banner(
        ws,
        "L",
        '="Points x PointRate (currently $"&TEXT(PointRate,"#,##0.00")'
        '&"/point) -- PointRate defaults to 0 (RATE_MISSING convention): visible zero, '
        'un-costed not free until a real rate is supplied."',
    )

    row = 3
    for c, h in enumerate(POD_COLS, start=1):
        ws.cell(row=row, column=c, value=h)
    ws.cell(row=row, column=13, value="Dense-arm mult. (0=off)")
    style_header_row(ws, row, 13)
    row += 1

    all_rows = POD_ROWS_A + POD_ROWS_B + POD_ROWS_C + POD_ROWS_D + POD_ROWS_E
    points_start = row
    for pod, title, components, points, kind, params, status, basis in all_rows:
        r0 = row
        ws.cell(row=r0, column=1, value=pod)
        ws.cell(row=r0, column=2, value=title)
        ws.cell(row=r0, column=3, value=components)
        pts_cell = ws.cell(row=r0, column=4, value=points)
        pts_cell.font = BLUE
        pts_cell.number_format = QTY_FMT

        if kind == "llm_dense_optional":
            # Amber DenseArmMultiplier input lives in column M (beyond Basis), 0 = off.
            amber_cell = ws.cell(row=r0, column=13, value=0)
            amber_cell.fill = AMBER_FILL
            amber_cell.font = BLUE
            amber_cell.number_format = QTY_FMT
            params = dict(params, _cell=f"$M${r0}")

        freq_txt, vol_txt, tin_txt, tout_txt, cost_f = _pod_cost_formula(
            kind, params, nr
        )
        ws.cell(row=r0, column=5, value=freq_txt)
        ws.cell(row=r0, column=6, value=vol_txt)
        ws.cell(row=r0, column=7, value=tin_txt)
        ws.cell(row=r0, column=8, value=tout_txt)

        cost_cell = ws.cell(row=r0, column=9, value=cost_f)
        cost_cell.font = BLACK
        cost_cell.number_format = MONEY2_FMT

        build_cell = ws.cell(row=r0, column=10, value=f"=D{r0}*PointRate")
        build_cell.font = BLACK
        build_cell.number_format = MONEY_FMT

        status_cell(ws, f"K{r0}", status)
        basis_cell = ws.cell(row=r0, column=12, value=basis)
        basis_cell.font = GREY
        basis_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r0].height = 30
        row += 1
    points_end = row - 1

    style_section(ws, row, "L", "TOTALS")
    row += 1
    ws.cell(row=row, column=2, value="Total points / monthly cost / build cost")
    tot_pts = ws.cell(row=row, column=4, value=f"=SUM(D{points_start}:D{points_end})")
    tot_pts.font = BLACK
    tot_pts.number_format = QTY_FMT
    tot_cost = ws.cell(row=row, column=9, value=f"=SUM(I{points_start}:I{points_end})")
    tot_cost.font = BLACK
    tot_cost.number_format = MONEY_FMT
    tot_build = ws.cell(
        row=row, column=10, value=f"=SUM(J{points_start}:J{points_end})"
    )
    tot_build.font = BLACK
    tot_build.number_format = MONEY_FMT
    for c in (4, 9, 10):
        ws.cell(row=row, column=c).fill = SUBTOTAL_FILL

    wb.defined_names["TotalBuildPoints"] = DefinedName(
        "TotalBuildPoints", attr_text=f"'Pod Build'!$D${row}"
    )
    wb.defined_names["TotalBuildCost"] = DefinedName(
        "TotalBuildCost", attr_text=f"'Pod Build'!$J${row}"
    )
    ws.freeze_panes = "A4"


TCO_COLS = ["Zone", "Component", "Monthly cost formula", "Status", "Basis / source"]


def build_monthly_tco(wb, rates, rate_cells, assumptions_log):
    nr = resolve_named_rates(rates, rate_cells)
    ws = wb.create_sheet("Monthly TCO")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([28, 40, 60, 14, 70], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    style_title(
        ws,
        "E",
        "Monthly TCO -- grouped by zone, formula-linked to Pod Build",
        height=18,
    )
    style_banner(
        ws,
        "E",
        '="Contingency uplift "&TEXT(ContingencyPct,"0.0%")&" applied to the subtotal '
        "below. Three scenario rows show the LLM token cost under different framings -- "
        'only the CODE-VERIFIED LIKELY row (Control Panel defaults) feeds the total."',
    )

    row = 3
    for c, h in enumerate(TCO_COLS, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(TCO_COLS))
    row += 1

    sku_in = sku_gate("ModelSKU", nr["opus_in"], nr["sonnet_in"], nr["haiku_in"])
    sku_out = sku_gate("ModelSKU", nr["opus_out"], nr["sonnet_out"], nr["haiku_out"])

    def add(zone, name, formula, status, basis):
        nonlocal row
        ws.cell(row=row, column=1, value=zone)
        ws.cell(row=row, column=2, value=name)
        fc = ws.cell(row=row, column=3, value=formula)
        fc.font = GREEN if "'Pod Build'!" in formula else BLACK
        fc.number_format = MONEY2_FMT
        status_cell(ws, f"D{row}", status)
        bc = ws.cell(row=row, column=5, value=basis)
        bc.font = GREY
        bc.alignment = Alignment(wrap_text=True, vertical="top")
        r = row
        row += 1
        return r

    style_section(ws, row, "E", "ZONE 1 -- INGESTION")
    row += 1
    zone1_row = add(
        "Zone1",
        "API GW + Lambda ETL + S3 + SQS + EventBridge (Pod A total)",
        "='Pod Build'!$I$4+'Pod Build'!$I$5+'Pod Build'!$I$6+'Pod Build'!$I$7"
        "+'Pod Build'!$I$8+'Pod Build'!$I$9+'Pod Build'!$I$10+'Pod Build'!$I$11"
        "+'Pod Build'!$I$12+'Pod Build'!$I$13+'Pod Build'!$I$14+'Pod Build'!$I$15",
        "DERIVED",
        "Cross-links to Pod Build Pod A rows (4-15); avoids re-deriving the "
        "same formulas twice.",
    )

    style_section(ws, row, "E", "ZONE 2 -- ETL / VALIDATION")
    row += 1
    zone2_row = add(
        "Zone2",
        "Sanity check + quarantine (folded into Pod A compute)",
        "=0",
        "DERIVED",
        "No incremental Zone2 compute beyond Zone1's Ingestion Coordinator Lambda -- "
        "etl_handler.py validate/quarantine run in-process.",
    )

    style_section(ws, row, "E", "ZONE 3 -- MATCHING")
    row += 1
    zone3_row = add(
        "Zone3",
        "Matching Engine + Titan + OpenSearch + Aurora (Pod B total)",
        "='Pod Build'!$I$16+'Pod Build'!$I$17+'Pod Build'!$I$18+'Pod Build'!$I$19"
        "+'Pod Build'!$I$20+'Pod Build'!$I$21",
        "DERIVED",
        "Cross-links to Pod Build Pod B rows (16-21).",
    )

    style_section(ws, row, "E", "ZONE 4 -- AGENTIC")
    row += 1
    orch_row = add(
        "Zone4",
        "Strands Orchestrator + AgentCore Memory (Pod C infra rows)",
        "='Pod Build'!$I$22+'Pod Build'!$I$23+'Pod Build'!$I$24+'Pod Build'!$I$25",
        "DERIVED",
        "Cross-links to Pod Build Pod C rows 22-25 (orchestrator Lambda + "
        "AgentCore Memory short-term/long-term/retrieval).",
    )
    likely_row = add(
        "Zone4",
        "LLM tokens -- CODE-VERIFIED LIKELY (default, feeds total)",
        "='Pod Build'!$I$26+'Pod Build'!$I$27",
        "VERIFIED",
        "Sum of Specialist + Verifier lines (Pod Build rows 26-27) at the "
        "Control Panel's CODE-VERIFIED defaults (0.8K in / 0.25K out / 2 calls-item). "
        "llm_call_model.md SS3: ~$203/mo at Opus 4.8, 9,890 items/mo.",
    )
    high_row = add(
        "Zone4",
        "LLM tokens -- CLIENT-SKETCH HIGH (scenario only, NOT in total)",
        f"=LlmItems*4*(50/1000*{sku_in}+1/1000*{sku_out})",
        "ASSUMPTION",
        "Client's original sketch: 4 calls/item, 50K in / 1K out per call. "
        "Shown as a labelled scenario for comparison -- excluded from MonthlyTCO total. "
        "llm_call_model.md SS3 narrative.",
    )
    dense_row = add(
        "Zone4",
        "LLM tokens -- DENSE-ARM OPTIONAL (default OFF, scenario only)",
        "='Pod Build'!$I$28",
        "DERIVED",
        "Cross-links to Pod Build row 28 (opus_dense per taxonomy node); "
        "amber multiplier defaults 0 (off), matching deployed SCUDO_DENSE_BACKEND="
        "jaro_winkler. llm_call_model.md SS1/SS3 (14-node demo fixture if enabled).",
    )
    eval_row = add(
        "Zone4",
        "Bedrock Evaluations (judge -- folded into Specialist/Verifier tokens)",
        "='Pod Build'!$I$29",
        "ASSUMPTION",
        "No separate meter; $0 incremental (Pod Build row 29).",
    )

    style_section(ws, row, "E", "ZONE 5 -- PERSISTENCE / HITL")
    row += 1
    zone5_row = add(
        "Zone5",
        "Outbox + AppSync + ECS + projection sink (Pod D total)",
        "='Pod Build'!$I$30+'Pod Build'!$I$31+'Pod Build'!$I$32+'Pod Build'!$I$33"
        "+'Pod Build'!$I$34",
        "DERIVED",
        "Cross-links to Pod Build Pod D rows 30-34.",
    )

    style_section(ws, row, "E", "PLATFORM -- ALWAYS-ON (POD E)")
    row += 1
    platform_row = add(
        "Platform",
        "FalkorDB + CloudFront + ALB + NAT + Secrets + VPC endpoints + CodeBuild",
        "='Pod Build'!$I$35+'Pod Build'!$I$36+'Pod Build'!$I$37+'Pod Build'!$I$38"
        "+'Pod Build'!$I$39+'Pod Build'!$I$40+'Pod Build'!$I$41",
        "DERIVED",
        "Cross-links to Pod Build Pod E rows 35-41.",
    )

    style_section(ws, row, "E", "TOTALS")
    row += 1
    subtotal_row = row
    included_rows = [
        zone1_row,  # Zone1 Pod A
        zone2_row,  # Zone2 (=0)
        zone3_row,  # Zone3 Pod B (Matching Engine + Titan + OpenSearch + Aurora)
        orch_row,  # Zone4 orchestrator + AgentCore Memory
        likely_row,  # Zone4 LIKELY LLM tokens (default in total)
        dense_row,  # Zone4 dense-arm optional (0 by default, safe to include)
        eval_row,  # Zone4 evaluations ($0)
        zone5_row,  # Zone5 Pod D
        platform_row,  # Platform Pod E
    ]
    ws.cell(row=row, column=2, value="Subtotal (excludes HIGH scenario row)")
    sub_cell = ws.cell(
        row=row,
        column=3,
        value="=" + "+".join(f"C{r}" for r in included_rows),
    )
    sub_cell.font = BLACK
    sub_cell.number_format = MONEY2_FMT
    ws.cell(row=row, column=1).fill = SUBTOTAL_FILL
    ws.cell(row=row, column=3).fill = SUBTOTAL_FILL
    row += 1

    contingency_row = row
    ws.cell(row=row, column=2, value="Contingency uplift")
    cont_cell = ws.cell(row=row, column=3, value=f"=C{subtotal_row}*ContingencyPct")
    cont_cell.font = BLACK
    cont_cell.number_format = MONEY2_FMT
    row += 1

    total_row = row
    ws.cell(row=row, column=2, value="MONTHLY TCO (at Control Panel defaults)")
    total_cell = ws.cell(
        row=row, column=3, value=f"=C{subtotal_row}+C{contingency_row}"
    )
    total_cell.font = Font(bold=True)
    total_cell.number_format = MONEY_FMT
    ws.cell(row=row, column=1).fill = SUBTOTAL_FILL
    ws.cell(row=row, column=3).fill = SUBTOTAL_FILL

    wb.defined_names["MonthlyTCO"] = DefinedName(
        "MonthlyTCO", attr_text=f"'Monthly TCO'!$C${total_row}"
    )
    ws.freeze_panes = "A4"


MONTH_LABELS = [
    "Nov 2026",
    "Dec 2026",
    "Jan 2027",
    "Feb 2027",
    "Mar 2027",
    "Apr 2027",
    "May 2027",
    "Jun 2027",
    "Jul 2027",
    "Aug 2027",
    "Sep 2027",
    "Oct 2027",
]


def build_12mo_run_cost(wb, rates, rate_cells):
    """Months 1-12 (Nov 2026 -> Oct 2027), ramp profile M1 50%/M2 75%/M3+ 100% on
    volume-driven Pod Build lines; always-on infra lines carry flat cost every
    month (no ramp). One row per Pod Build cost line, cross-linked via
    ='Pod Build'!$I$<row> (green font = cross-sheet link)."""
    nr = resolve_named_rates(rates, rate_cells)
    ws = wb.create_sheet("12mo Run Cost")
    ws.sheet_view.showGridLines = False
    widths = [40, 12] + [11] * 12 + [14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_col = get_column_letter(2 + 12 + 1)  # A,B + 12 months + total = col 15 = O
    style_title(
        ws, last_col, "12-mo Run Cost -- ramp profile, Nov 2026 go-live", height=18
    )
    style_banner(
        ws,
        last_col,
        '="Ramp: M1 "&TEXT(RampM1,"0%")&" / M2 "&TEXT(RampM2,"0%")&" / M3+ '
        '"&TEXT(RampM3Plus,"0%")&" of steady-state volume -- applies to volume-driven '
        "meters only. Always-on infra (Aurora, OpenSearch, ECS, NAT, ALB, endpoints) "
        'does not ramp."',
    )

    row = 3
    ws.cell(row=row, column=1, value="Component")
    ws.cell(row=row, column=2, value="Ramp class")
    for m, label in enumerate(MONTH_LABELS, start=3):
        ws.cell(row=row, column=m, value=label)
    ws.cell(row=row, column=15, value="12-mo Total")
    style_header_row(ws, row, 15)
    row += 1

    ramp_row = row
    ws.cell(row=row, column=1, value="Ramp % (of steady-state) -- editable")
    m1_cell = ws.cell(row=row, column=3, value=0.5)
    m2_cell = ws.cell(row=row, column=4, value=0.75)
    m1_cell.fill = AMBER_FILL
    m1_cell.font = BLUE
    m1_cell.number_format = PCT_FMT
    m2_cell.fill = AMBER_FILL
    m2_cell.font = BLUE
    m2_cell.number_format = PCT_FMT
    m3_col = 5
    m3_cell = ws.cell(row=row, column=m3_col, value=1.0)
    m3_cell.fill = AMBER_FILL
    m3_cell.font = BLUE
    m3_cell.number_format = PCT_FMT
    for m in range(m3_col + 1, 15):
        c = ws.cell(row=row, column=m, value=f"={get_column_letter(m3_col)}${row}")
        c.font = BLACK
        c.number_format = PCT_FMT
    wb.defined_names["RampM1"] = DefinedName(
        "RampM1", attr_text=f"'12mo Run Cost'!$C${row}"
    )
    wb.defined_names["RampM2"] = DefinedName(
        "RampM2", attr_text=f"'12mo Run Cost'!$D${row}"
    )
    wb.defined_names["RampM3Plus"] = DefinedName(
        "RampM3Plus", attr_text=f"'12mo Run Cost'!${get_column_letter(m3_col)}${row}"
    )
    row += 1

    all_rows = POD_ROWS_A + POD_ROWS_B + POD_ROWS_C + POD_ROWS_D + POD_ROWS_E
    comp_start = row
    for i, (pod, title, components, points, kind, params, status, basis) in enumerate(
        all_rows, start=4
    ):
        p = dict(params)
        if kind == "llm_dense_optional":
            p["_cell"] = f"'Pod Build'!$M${i}"
        freq_txt = _pod_cost_formula(kind, p, nr)[0]
        ramps = freq_txt != "Always-on"
        ws.cell(row=row, column=1, value=f"Pod {pod}: {title}")
        ws.cell(row=row, column=2, value="Ramps" if ramps else "Always-on")
        for m in range(12):
            col = 3 + m
            ramp_col_letter = get_column_letter(
                3 if m == 0 else (4 if m == 1 else m3_col)
            )
            if ramps:
                f = f"='Pod Build'!$I${i}*{ramp_col_letter}${ramp_row}"
            else:
                f = f"='Pod Build'!$I${i}"
            cell = ws.cell(row=row, column=col, value=f)
            cell.font = GREEN
            cell.number_format = MONEY2_FMT
        total_cell = ws.cell(
            row=row,
            column=15,
            value=f"=SUM(C{row}:N{row})",
        )
        total_cell.font = BLACK
        total_cell.number_format = MONEY_FMT
        row += 1
    comp_end = row - 1

    style_section(ws, row, last_col, "TOTALS")
    row += 1
    totals_row = row
    ws.cell(row=row, column=1, value="Monthly total (all Pod Build lines)")
    for m in range(12):
        col = 3 + m
        col_l = get_column_letter(col)
        c = ws.cell(
            row=row, column=col, value=f"=SUM({col_l}{comp_start}:{col_l}{comp_end})"
        )
        c.font = BLACK
        c.number_format = MONEY_FMT
        c.fill = SUBTOTAL_FILL
    run12mo_cell = ws.cell(row=row, column=15, value=f"=SUM(C{row}:N{row})")
    run12mo_cell.font = Font(bold=True)
    run12mo_cell.number_format = MONEY_FMT
    run12mo_cell.fill = SUBTOTAL_FILL
    ws.cell(row=row, column=1).fill = SUBTOTAL_FILL

    wb.defined_names["Run12mo"] = DefinedName(
        "Run12mo", attr_text=f"'12mo Run Cost'!$O${totals_row}"
    )
    ws.freeze_panes = "C4"


# Gantt: 14-week core (kickoff w/c 2026-07-27 -> go-live 2026-11-01), weeks 1-14
# one column each starting at column L (col 12), slip headroom to week 18 (col 16).
GANTT_WEEK0_COL = 12  # column L = week 1
GANTT_LAST_WEEK_COL = GANTT_WEEK0_COL + 17  # allow to week 18
GANTT_CORE_WEEKS = 14
# month boundaries (first week-column of each new month, computed from kickoff date)
GANTT_MONTH_LABELS = [
    (1, "Jul 26"),
    (2, "Aug 26"),
    (6, "Sep 26"),
    (10, "Oct 26"),
    (14, "Nov 26"),
]

# (pod, id, task, is_parent, start_wk, wks, basis)
GANTT_ROWS = [
    (
        "A",
        "G-A-1",
        "Pod A -- Ingestion foundation",
        True,
        1,
        2,
        "API GW, S3 buckets, EventBridge rule/scheduler, SQS provisioned first (shared foundation).",
    ),
    (
        "A",
        "G-A-1a",
        "    - IaC: template.yaml Pod A resources",
        False,
        1,
        1,
        "SAM stack apply.",
    ),
    (
        "A",
        "G-A-1b",
        "    - Smoke test: upload -> canonical S3",
        False,
        2,
        1,
        "End-to-end ingest smoke.",
    ),
    (
        "A",
        "G-A-2",
        "Pod A -- ETL worker + validation live",
        True,
        3,
        3,
        "Ingestion Coordinator Lambda + sanity-check/quarantine path hardened.",
    ),
    (
        "A",
        "G-A-2a",
        "    - Build EtlFn + DLQ wiring",
        False,
        3,
        2,
        "template.yaml EtlFn.",
    ),
    (
        "A",
        "G-A-2b",
        "    - Quarantine + CloudWatch dashboards",
        False,
        5,
        1,
        "arch_infra_persistence.md A.8.",
    ),
    (
        "B",
        "G-B-1",
        "Pod B -- Matching engine + Titan + OpenSearch",
        True,
        3,
        4,
        "AIA Matching Engine gates A/B/C, Titan embeddings write path, OpenSearch index.",
    ),
    (
        "B",
        "G-B-1a",
        "    - Aurora 2xlarge provisioned",
        False,
        3,
        1,
        "System-of-record instance.",
    ),
    (
        "B",
        "G-B-1b",
        "    - OpenSearch fuzzy+kNN index build",
        False,
        4,
        2,
        "Dense + sparse arms.",
    ),
    (
        "B",
        "G-B-1c",
        "    - Matching gate A/B/C integration test",
        False,
        6,
        1,
        "Confidence-band gate.",
    ),
    (
        "C",
        "G-C-1",
        "Pod C -- Strands orchestrator + agents",
        True,
        3,
        5,
        "Orchestrator, AgentCore Memory, specialist + verifier agents wired to Bedrock.",
    ),
    (
        "C",
        "G-C-1a",
        "    - Strands Orchestrator Opus 4.8 wiring",
        False,
        3,
        2,
        "orchestrator.py.",
    ),
    (
        "C",
        "G-C-1b",
        "    - AgentCore Memory short/long-term",
        False,
        5,
        2,
        "Events + records + retrieval.",
    ),
    (
        "C",
        "G-C-1c",
        "    - Specialist + Verifier + Evaluations",
        False,
        7,
        1,
        "10-dim verifier gate.",
    ),
    (
        "D",
        "G-D-1",
        "Pod D -- Persistence + HITL",
        True,
        3,
        5,
        "Outbox writer, sweep schedule, AppSync HITL API, ECS reviewer console.",
    ),
    (
        "D",
        "G-D-1a",
        "    - JAPI Persist outbox + sweep",
        False,
        3,
        2,
        "Transactional outbox pattern.",
    ),
    (
        "D",
        "G-D-1b",
        "    - HITL AppSync + ECS console",
        False,
        5,
        2,
        "Reviewer UI live.",
    ),
    (
        "D",
        "G-D-1c",
        "    - Projection sink + CloudFront invalidation",
        False,
        7,
        1,
        "S3 projection publish.",
    ),
    (
        "E",
        "G-E-1",
        "Pod E -- Platform (VPC, security, CI/CD)",
        True,
        1,
        3,
        "FalkorDB, CloudFront, ALB, NAT, Secrets Manager, VPC endpoints, CodeBuild.",
    ),
    (
        "E",
        "G-E-1a",
        "    - VPC endpoints + NAT + Secrets Manager",
        False,
        1,
        2,
        "Network foundation.",
    ),
    (
        "E",
        "G-E-1b",
        "    - CodeBuild pipeline + FalkorDB ECS",
        False,
        2,
        2,
        "CI/CD + retrieval store.",
    ),
    (
        "ALL",
        "G-X-1",
        "Integration -- first delivery E2E",
        True,
        8,
        2,
        "First vendor delivery run end-to-end through all 5 zones.",
    ),
    (
        "ALL",
        "G-X-1a",
        "    - E2E dry run against fixture delivery",
        False,
        8,
        1,
        "Smoke run.",
    ),
    (
        "ALL",
        "G-X-1b",
        "    - Fix findings from dry run",
        False,
        9,
        1,
        "Bug-fix buffer.",
    ),
    (
        "ALL",
        "G-X-2",
        "Go-live readiness -- HITL live + cutover",
        True,
        11,
        4,
        "HITL review live for reviewers, production cutover rehearsal, go-live.",
    ),
    ("ALL", "G-X-2a", "    - HITL reviewer UAT", False, 11, 2, "Reviewer sign-off."),
    (
        "ALL",
        "G-X-2b",
        "    - Production cutover rehearsal",
        False,
        13,
        1,
        "Dry-run cutover.",
    ),
    ("ALL", "G-X-2c", "    - Go-live 2026-11-01", False, 14, 1, "Live."),
]


def build_gantt(wb):
    ws = wb.create_sheet("Gantt")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False

    col_widths = {1: 10, 2: 42, 3: 10, 4: 8, 5: 60}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for c in range(6, GANTT_LAST_WEEK_COL + 1):
        ws.column_dimensions[get_column_letter(c)].width = 3.5

    last_col = get_column_letter(GANTT_LAST_WEEK_COL)
    style_title(
        ws,
        last_col,
        "Gantt -- 14-week core (Pod lanes, parallel after 2-week foundation)",
        height=18,
    )
    style_banner(
        ws,
        last_col,
        '="Kickoff w/c 2026-07-27 -> go-live 2026-11-01. SlipWeeks="&TEXT(SlipWeeks,"0")'
        '&" applied via proportional stretch inside the core span, additive shift beyond '
        "it. Bars = belt-and-braces cell value + conditional formatting (uniform green = "
        'proposed/recommended)."',
    )

    header_row = 3
    week_row = 4
    milestone_row = 5
    ws.cell(row=header_row, column=2, value="Task")
    for wk, label in GANTT_MONTH_LABELS:
        ws.cell(
            row=header_row, column=GANTT_WEEK0_COL + wk - 1, value=label
        ).font = Font(bold=True)

    ws.cell(row=week_row, column=2, value="Week #")
    for wk in range(1, GANTT_LAST_WEEK_COL - GANTT_WEEK0_COL + 2):
        ws.cell(row=week_row, column=GANTT_WEEK0_COL + wk - 1, value=wk)
    golive_col = GANTT_WEEK0_COL + GANTT_CORE_WEEKS - 1  # week 14 column
    bx = get_column_letter(GANTT_LAST_WEEK_COL + 2)  # helper cells past the grid
    by = get_column_letter(GANTT_LAST_WEEK_COL + 3)
    bz = get_column_letter(GANTT_LAST_WEEK_COL + 4)
    ws[f"{bx}{week_row}"] = f"={GANTT_CORE_WEEKS}+SlipWeeks"  # GanttCorrGoLive
    ws[f"{by}{week_row}"] = f"={GANTT_CORE_WEEKS - 2}+SlipWeeks"  # band start
    ws[f"{bz}{week_row}"] = f"={GANTT_CORE_WEEKS + 2}+SlipWeeks"  # band end
    wb.defined_names["GanttCorrGoLive"] = DefinedName(
        "GanttCorrGoLive", attr_text=f"Gantt!${bx}${week_row}"
    )

    ws.cell(row=milestone_row, column=2, value="Go-live range")
    for wk in range(1, GANTT_LAST_WEEK_COL - GANTT_WEEK0_COL + 2):
        col_l = get_column_letter(GANTT_WEEK0_COL + wk - 1)
        formula = (
            f'=IF({col_l}${week_row}=${bx}${week_row},"▲",'
            f'IF({col_l}${week_row}=${by}${week_row},"◁",'
            f'IF({col_l}${week_row}=${bz}${week_row},"▷",'
            f"IF(AND({col_l}${week_row}>${by}${week_row},{col_l}${week_row}<${bz}${week_row}),"
            f'"·",""))))'
        )
        ws.cell(row=milestone_row, column=GANTT_WEEK0_COL + wk - 1, value=formula)

    headers = [
        "Id",
        "Task",
        "Pod",
        "Start wk",
        "Wks",
        "Live start",
        "Live end",
        "Basis / source",
    ]
    hcol = {
        "id": 1,
        "task": 2,
        "pod": 3,
        "startwk": 4,
        "wks": 5,
        "live_start": GANTT_LAST_WEEK_COL
        + 2
        + 4,  # placed after helper cells, non-visual
        "live_end": GANTT_LAST_WEEK_COL + 2 + 5,
        "basis": GANTT_LAST_WEEK_COL + 2 + 6,
    }
    ws.cell(row=6, column=hcol["id"], value="Id")
    ws.cell(row=6, column=hcol["task"], value="Task")
    ws.cell(row=6, column=hcol["pod"], value="Pod")
    ws.cell(row=6, column=hcol["startwk"], value="Start wk")
    ws.cell(row=6, column=hcol["wks"], value="Wks")
    ws.cell(row=6, column=hcol["live_start"], value="Live start")
    ws.cell(row=6, column=hcol["live_end"], value="Live end")
    ws.cell(row=6, column=hcol["basis"], value="Basis / source")
    style_header_row(ws, 6, GANTT_LAST_WEEK_COL)

    row = 7
    core = GANTT_CORE_WEEKS
    green_fill = PatternFill("solid", fgColor="FF2E7D32")
    pale_fill = PatternFill("solid", fgColor="FFA1C5A3")

    def week_cols_range():
        return f"{get_column_letter(GANTT_WEEK0_COL)}:{get_column_letter(GANTT_LAST_WEEK_COL)}"

    parent_start_row = None
    child_block_start = None
    for pod, task_id, task, is_parent, start_wk, wks, basis in GANTT_ROWS:
        r = row
        ws.cell(row=r, column=hcol["id"], value=task_id)
        tcell = ws.cell(row=r, column=hcol["task"], value=task)
        tcell.font = Font(bold=is_parent)
        ws.cell(row=r, column=hcol["pod"], value=pod)
        sw_cell = ws.cell(row=r, column=hcol["startwk"], value=start_wk)
        sw_cell.font = BLUE
        wk_cell = ws.cell(row=r, column=hcol["wks"], value=wks)
        wk_cell.font = BLUE

        g, h = f"D{r}", f"E{r}"
        if is_parent:
            i_formula = (
                f"=MIN(18,IF({g}<={core},ROUND(({g}-1)*({core}+SlipWeeks)/{core},0)+1,"
                f"{g}+SlipWeeks))"
            )
            j_formula = (
                f"=MIN(18,IF({g}+{h}-1<={core},MAX(ROUND(({g}+{h}-1)*({core}+SlipWeeks)/{core},0),"
                f"F{r}),{g}+{h}-1+SlipWeeks))"
            )
            parent_i, parent_j = f"F{r}", f"G{r}"
            parent_start_row = r
        else:
            pi, pj = f"F{parent_start_row}", f"G{parent_start_row}"
            pg, ph = f"D{parent_start_row}", f"E{parent_start_row}"
            i_formula = (
                f"=MEDIAN(${pi},${pj},${pi}+ROUND(({g}-${pg})/${ph}*(${pj}-${pi}+1),0))"
            )
            j_formula = f"=MEDIAN(F{r},${pj},${pi}+ROUND(({g}+{h}-${pg})/${ph}*(${pj}-${pi}+1),0)-1)"
        ws.cell(row=r, column=6, value=i_formula).font = BLACK
        ws.cell(row=r, column=7, value=j_formula).font = BLACK

        for wk in range(1, GANTT_LAST_WEEK_COL - GANTT_WEEK0_COL + 2):
            col = GANTT_WEEK0_COL + wk - 1
            col_l = get_column_letter(col)
            bar_formula = (
                f'=IF(AND({col_l}${week_row}>=$F{r},{col_l}${week_row}<=$G{r}),1,"")'
            )
            bar_cell = ws.cell(row=r, column=col, value=bar_formula)
            bar_cell.number_format = HIDE_FMT

        basis_cell = ws.cell(row=r, column=hcol["basis"], value=basis)
        basis_cell.font = GREY
        basis_cell.alignment = Alignment(wrap_text=True, vertical="top")

        rng = f"{get_column_letter(GANTT_WEEK0_COL)}{r}:{get_column_letter(GANTT_LAST_WEEK_COL)}{r}"
        cf_formula = f"AND({get_column_letter(GANTT_WEEK0_COL)}${week_row}>=$F{r},{get_column_letter(GANTT_WEEK0_COL)}${week_row}<=$G{r})"
        rule = FormulaRule(
            formula=[cf_formula], fill=green_fill if is_parent else pale_fill
        )
        ws.conditional_formatting.add(rng, rule)

        if not is_parent:
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = True
        row += 1

    ws.freeze_panes = f"{get_column_letter(GANTT_WEEK0_COL)}{6}"


# Monthly TCO rows feeding the Overview waterfall (row, label) -- excludes the
# HIGH scenario row (13) and section headers, matches Monthly TCO's own subtotal.
OVERVIEW_WATERFALL_ROWS = [
    (5, "Zone1 Ingestion"),
    (7, "Zone2 ETL/Validation"),
    (9, "Zone3 Matching"),
    (11, "Zone4 Orchestrator+Memory"),
    (12, "Zone4 LLM tokens (LIKELY)"),
    (14, "Zone4 Dense-arm (opt, 0 default)"),
    (15, "Zone4 Evaluations"),
    (17, "Zone5 Persistence/HITL"),
    (19, "Platform (always-on)"),
]


def build_overview(wb):
    ws = wb.create_sheet("Overview")
    ws.sheet_view.showGridLines = False
    widths = [26, 20, 20, 16, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    style_title(ws, "F", "Overview -- headline TCO + cost waterfall", height=18)
    style_banner(
        ws,
        "F",
        '="Monthly TCO $"&TEXT(MonthlyTCO,"#,##0.00")&" at Control Panel defaults '
        '("&ModelSKU&", "&TEXT(ItemsPerMonth,"#,##0")&" items/mo). 12-mo run cost '
        '$"&TEXT(Run12mo,"#,##0")&". Build: "&TEXT(TotalBuildPoints,"#,##0")&" points x '
        '$"&TEXT(PointRate,"#,##0.00")&"/point = $"&TEXT(TotalBuildCost,"#,##0")&"."',
    )

    row = 3
    style_section(ws, row, "F", "HEADLINE")
    row += 1
    headline_cell = ws.cell(
        row=row, column=2, value="Monthly TCO (at Control Panel defaults)"
    )
    v_cell = ws.cell(row=row, column=3, value="=MonthlyTCO")
    v_cell.font = GREEN
    v_cell.number_format = MONEY_FMT
    row += 1
    ws.cell(row=row, column=2, value="12-month run cost total")
    r12 = ws.cell(row=row, column=3, value="=Run12mo")
    r12.font = GREEN
    r12.number_format = MONEY_FMT
    row += 1
    ws.cell(row=row, column=2, value="Total build points")
    tbp = ws.cell(row=row, column=3, value="=TotalBuildPoints")
    tbp.font = GREEN
    tbp.number_format = QTY_FMT
    row += 1
    ws.cell(row=row, column=2, value="Total build cost (Points x PointRate)")
    tbc = ws.cell(row=row, column=3, value="=TotalBuildCost")
    tbc.font = GREEN
    tbc.number_format = MONEY_FMT
    row += 2

    style_section(ws, row, "F", "COST WATERFALL -- MONTHLY TCO BY ZONE")
    row += 1
    hdr_row = row
    for c, h in enumerate(
        ["Zone / component", "Base (invisible)", "Block (visible)", "Delta"], start=1
    ):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, 4)
    row += 1
    wf_start = row
    prev_cum_cell = None
    for i, (mtco_row, label) in enumerate(OVERVIEW_WATERFALL_ROWS):
        r = row
        ws.cell(row=r, column=1, value=label)
        delta_cell = ws.cell(row=r, column=4, value=f"='Monthly TCO'!$C${mtco_row}")
        delta_cell.font = GREEN
        delta_cell.number_format = MONEY2_FMT
        if prev_cum_cell is None:
            base_formula = "=0"
        else:
            base_formula = f"=E{r - 1}"
        base_cell = ws.cell(row=r, column=2, value=base_formula)
        base_cell.font = BLACK
        base_cell.number_format = MONEY2_FMT
        block_cell = ws.cell(row=r, column=3, value=f"=ABS(D{r})")
        block_cell.font = BLACK
        block_cell.number_format = MONEY2_FMT
        cum_cell = ws.cell(row=r, column=5, value=f"=B{r}+D{r}")
        cum_cell.font = BLACK
        cum_cell.number_format = MONEY2_FMT
        prev_cum_cell = cum_cell
        row += 1
    wf_end = row - 1

    row += 1
    ws.cell(
        row=row,
        column=1,
        value="Cumulative total (should equal Subtotal on Monthly TCO)",
    )
    tot_cell = ws.cell(row=row, column=5, value=f"=E{wf_end}")
    tot_cell.font = Font(bold=True)
    tot_cell.number_format = MONEY_FMT
    row += 2

    style_section(ws, row, "F", "TOP COST DRIVERS")
    row += 1
    driver_hdr = row
    ws.cell(row=row, column=1, value="Rank")
    ws.cell(row=row, column=2, value="Zone / component")
    ws.cell(row=row, column=3, value="Monthly cost")
    style_header_row(ws, row, 3)
    row += 1
    for rank in range(1, 4):
        ws.cell(row=row, column=1, value=rank)
        large_fn = "LARGE" if rank == 1 else "LARGE"
        rank_cell = ws.cell(
            row=row,
            column=3,
            value=f"={large_fn}($D${wf_start}:$D${wf_end},{rank})",
        )
        rank_cell.font = BLACK
        rank_cell.number_format = MONEY2_FMT
        name_cell = ws.cell(
            row=row,
            column=2,
            value=f"=INDEX($A${wf_start}:$A${wf_end},MATCH(C{row},$D${wf_start}:$D${wf_end},0))",
        )
        name_cell.font = BLACK
        row += 1

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Monthly TCO waterfall by zone"
    base_ref = Reference(ws, min_col=2, min_row=hdr_row, max_row=wf_end)
    block_ref = Reference(ws, min_col=3, min_row=hdr_row, max_row=wf_end)
    cats_ref = Reference(ws, min_col=1, min_row=wf_start, max_row=wf_end)
    chart.add_data(base_ref, titles_from_data=True)
    chart.add_data(block_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    base_series = chart.series[0]
    base_series.graphicalProperties.noFill = True
    chart.series[1].graphicalProperties.solidFill = "2E7D32"
    chart.height = 9
    chart.width = 20
    ws.add_chart(chart, f"H{hdr_row}")

    ws.freeze_panes = "A3"


ASSUMPTIONS_EXTRA = [
    (
        "Control Panel",
        "AuroraSize",
        "AuroraSize=2xlarge vs Serverless v2",
        "ASSUMPTION",
        "Client sketch specifies a fixed db.r6g.2xlarge Single-AZ instance; the deployed "
        "repo stack (infra/scudo-poc-app.yaml) actually provisions Aurora Serverless v2 "
        "with an ACU range, not a fixed instance class. The workbook models the client's "
        "stated instance class (fixed-price, easier to reason about); to correct, re-price "
        "against the Serverless v2 ACU-hour meters in Rates and swap the AuroraSize "
        "IF-gate in Pod Build row 21 for an ACU-hour formula.",
    ),
    (
        "Control Panel",
        "StoragePerItemKB",
        "StoragePerItemKB=50 (canonical JSON-LD per product)",
        "ASSUMPTION",
        "No repo-measured average canonical-JSON-LD payload size was found; 50KB/item is a "
        "round-number placeholder. To correct: sample real canonical objects from the "
        "CleanCanonicalBucket and replace the Control Panel default with a measured median.",
    ),
    (
        "Pod Build row 5",
        "Lambda authorizer",
        "No JWT/Lambda authorizer exists in the deployed stack",
        "RATE_MISSING",
        "Auth today is an API-key header (arch_infra_persistence.md D.1), not a Lambda "
        "authorizer. Carried at $0 (un-costed, not free) until this component is actually "
        "built; if built, price it as a per-request Lambda invocation on the api_gw route.",
    ),
    (
        "Pod Build row 17",
        "AIA Matching Engine gate B (precedent)",
        "No RDS Data API / ExecuteStatement per-request meter in Rates",
        "RATE_MISSING",
        "research/rates.json has zero RDS Data API entries -- only Aurora instance-hour and "
        "storage/IO meters exist. The underlying Aurora instance cost IS priced (Pod Build "
        "row 21); only the per-call Data API charge itself is un-costed. To correct: source "
        "an official RDS Data API ExecuteStatement rate and add it to Rates.",
    ),
    (
        "Pod Build row 30",
        "JAPI Persist outbox writer",
        "Same RDS Data API rate gap as row 17",
        "RATE_MISSING",
        "~10 ExecuteStatement calls/product is VERIFIED (arch_infra_persistence.md), but the "
        "per-call rate itself is RATE_MISSING for the same reason as row 17 above.",
    ),
    (
        "Pod Build row 41",
        "CodeBuild (console/dashboard build)",
        "No CodeBuild meter in Rates",
        "RATE_MISSING",
        "research/rates.json has zero CodeBuild entries (no build-minute meter sourced). "
        "Carried at $0 (un-costed, not free) until a CodeBuild compute-type rate is added "
        "to Rates.",
    ),
    (
        "Pod Build row 28 / Monthly TCO Zone4",
        "Dense-arm optional (opus_dense per taxonomy node)",
        "Off by default, matches deployed config",
        "DERIVED",
        "Deployed SCUDO_DENSE_BACKEND=jaro_winkler (infra/scudo-poc-app.yaml:76-79), not "
        "opus -- the dense-arm LLM-scoring path is currently disabled in production. The "
        "amber multiplier on Pod Build row 28 defaults to 0 for this reason; set >0 only "
        "to model a future opus_dense_score rollout.",
    ),
    (
        "Monthly TCO Zone4",
        "LLM tokens -- three scenario framings",
        "LIKELY (feeds total) vs HIGH (scenario only) vs DENSE (off)",
        "VERIFIED / ASSUMPTION",
        "LIKELY is CODE-VERIFIED against llm_call_model.md (~19.8K calls/mo, ~15.8M in + "
        "4.9M out tokens, ~$203/mo at Opus 4.8 defaults). HIGH replays the client's original "
        "sketch numbers (4 calls/item, 50K in / 1K out) as a labelled comparison row, NOT "
        "fed into the total -- do not confuse the two when quoting a headline figure.",
    ),
    (
        "Pod Build (Points column)",
        "Story points -- ASSUMPTION rows",
        "Client-supplied points kept where given; ASSUMPTION points elsewhere",
        "ASSUMPTION",
        "Rows without an explicit client point estimate use relative-sizing placeholders "
        "(blue font, hardcoded). PointRate defaults to 0 by design (RATE_MISSING "
        "convention) so the Build cost column shows visible zeros, not silent omission, "
        "until JPMC supplies a $/point rate.",
    ),
]


def build_assumptions(wb, assumptions_log=None):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    widths = [22, 34, 40, 14, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    style_title(
        ws,
        "E",
        "Assumptions -- every ASSUMPTION / RATE_MISSING line, restated",
        height=18,
    )
    style_banner(
        ws,
        "E",
        '="This sheet is a static restatement (not formula-linked) of every non-VERIFIED '
        "line elsewhere in the workbook, so a reviewer can find every place a number was "
        'estimated rather than sourced, and what to do to correct it."',
    )

    row = 3
    for c, h in enumerate(
        ["Location", "Item", "What was assumed", "Status", "Basis / how to correct"],
        start=1,
    ):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, 5)
    row += 1

    all_pod_rows = POD_ROWS_A + POD_ROWS_B + POD_ROWS_C + POD_ROWS_D + POD_ROWS_E
    for i, (pod, title, components, points, kind, params, status, basis) in enumerate(
        all_pod_rows, start=4
    ):
        if status not in ("ASSUMPTION", "RATE_MISSING", "INDICATIVE"):
            continue
        ws.cell(row=row, column=1, value=f"Pod Build row {i}")
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value=components)
        status_cell(ws, f"D{row}", status)
        bc = ws.cell(row=row, column=5, value=basis)
        bc.font = GREY
        bc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 28
        row += 1

    row += 1
    style_section(
        ws, row, "E", "OTHER NOTES (cross-cutting, not tied to a single Pod Build row)"
    )
    row += 1
    for location, item, what, status, basis in ASSUMPTIONS_EXTRA:
        ws.cell(row=row, column=1, value=location)
        ws.cell(row=row, column=2, value=item)
        ws.cell(row=row, column=3, value=what)
        status_cell(ws, f"D{row}", status)
        bc = ws.cell(row=row, column=5, value=basis)
        bc.font = GREY
        bc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    ws.freeze_panes = "A4"


def main():
    wb = Workbook()
    wb.remove(wb.active)

    rates = load_rates()
    assumptions_log = []  # (sheet, cell, label, status, basis)

    # build order (dependency-first); tab order fixed at the end
    build_control_panel(wb)
    rate_cells = build_rates(wb, rates)
    build_pod_build(wb, rates, rate_cells, assumptions_log)
    build_monthly_tco(wb, rates, rate_cells, assumptions_log)
    build_12mo_run_cost(wb, rates, rate_cells)
    build_gantt(wb)
    build_overview(wb)
    build_assumptions(wb, assumptions_log)

    order = [
        "Overview",
        "Control Panel >>>",
        "Rates",
        "Pod Build",
        "Monthly TCO",
        "12mo Run Cost",
        "Gantt",
        "Assumptions",
    ]
    wb._sheets = [wb[name] for name in order]
    for i, name in enumerate(order):
        wb[name].sheet_view.tabSelected = i == 0

    out = os.path.join(HERE, "JPMC_SCUDO_POAP_TCO_v1.xlsx")
    wb.save(out)
    print(f"Built {out}")


if __name__ == "__main__":
    main()
