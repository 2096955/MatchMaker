#!/usr/bin/env python3
"""Independent verification of JPMC_SCUDO_POAP_TCO_v1.xlsx.

Run AFTER recalc.py has produced cached values. Checks:
  1. contamination scrub (segro/hcl/azure/gbp/uksouth + pound sign) over values AND formulas
  2. five spot-checks against design-doc expectations (data_only values)
  3. structural checks: sheets, named ranges, formula density, status labels
Exit 0 = all pass; prints a findings report either way.
"""

import re
import sys
from openpyxl import load_workbook

PATH = (
    "/Users/anthonylui/MatchMaker/MatchMaker/jpmc-costings/JPMC_SCUDO_POAP_TCO_v1.xlsx"
)
SCRUB = re.compile(r"segro|hcl|azure|gbp|uksouth|£", re.IGNORECASE)
EXPECTED_SHEETS = [
    "Overview",
    "Control Panel >>>",
    "Rates",
    "Pod Build",
    "Monthly TCO",
    "12mo Run Cost",
    "Gantt",
    "Assumptions",
]
EXPECTED_NAMES = [
    "DeliveriesPerMonth",
    "ProductsPerDelivery",
    "DatasetCount",
    "LlmReachablePct",
    "ModelSKU",
    "InputTokensPerCallK",
    "OutputTokensPerCallK",
    "CallsPerItem",
    "PointRate",
    "EnvMultiplier",
    "AuroraSize",
    "HitlPct",
    "StoragePerItemKB",
    "SlipWeeks",
    "ContingencyPct",
    "ItemsPerMonth",
]

findings = []


def check(ok, label, detail=""):
    findings.append((ok, label, detail))
    print(f"{'PASS' if ok else 'FAIL'}  {label}" + (f"  -- {detail}" if detail else ""))


def main():
    wbf = load_workbook(PATH)  # formulas view
    wbv = load_workbook(PATH, data_only=True)  # cached values view

    # 1. contamination scrub over both views
    hits = []
    for wb, tag in ((wbf, "formula"), (wbv, "value")):
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and SCRUB.search(c.value):
                        hits.append(f"{tag}:{ws.title}!{c.coordinate}={c.value[:60]!r}")
    for name, dn in wbf.defined_names.items():
        blob = name + " " + str(dn.attr_text)
        if SCRUB.search(blob):
            hits.append(f"name:{blob[:60]!r}")
    for title in wbf.sheetnames:
        if SCRUB.search(title):
            hits.append(f"sheet:{title}")
    check(
        len(hits) == 0,
        "scrub segro/hcl/azure/gbp/uksouth/£",
        f"{len(hits)} hits" + ("; " + "; ".join(hits[:5]) if hits else ""),
    )

    # 2. structural: sheets present
    missing = [s for s in EXPECTED_SHEETS if s not in wbf.sheetnames]
    check(
        not missing,
        "8 expected sheets present",
        f"missing: {missing}" if missing else f"sheets: {wbf.sheetnames}",
    )

    # named ranges
    missing_n = [n for n in EXPECTED_NAMES if n not in wbf.defined_names]
    check(
        not missing_n,
        "core named ranges defined",
        f"missing: {missing_n}" if missing_n else f"{len(wbf.defined_names)} names",
    )

    # formula density: computed cells must be formulas, not constants
    n_formula = sum(
        1
        for ws in wbf.worksheets
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    )
    check(n_formula > 100, "formula count sane (>100)", f"{n_formula} formulas")

    # status labels present somewhere
    statuses = set()
    for ws in wbf.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip() in (
                    "VERIFIED",
                    "INDICATIVE",
                    "ASSUMPTION",
                    "DERIVED",
                    "RATE_MISSING",
                ):
                    statuses.add(c.value.strip())
    check(len(statuses) >= 3, "status vocabulary used", f"seen: {sorted(statuses)}")

    # residual formula errors in cached values (belt & braces after recalc.py)
    errs = []
    for ws in wbv.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if (
                    isinstance(c.value, str)
                    and c.value.startswith("#")
                    and c.value.endswith(("!", "?", "A"))
                ):
                    errs.append(f"{ws.title}!{c.coordinate}={c.value}")
    check(
        len(errs) == 0,
        "no #REF!/#DIV/0!/#VALUE!/#NAME?/#N/A in cached values",
        f"{len(errs)} errors" + ("; " + "; ".join(errs[:5]) if errs else ""),
    )

    run_spot_checks(wbf, wbv)

    n_fail = sum(1 for ok, *_ in findings if not ok)
    print(
        f"\n{'ALL CHECKS PASSED' if n_fail == 0 else str(n_fail) + ' CHECK(S) FAILED'} ({len(findings)} total)"
    )
    sys.exit(1 if n_fail else 0)


def resolve_name(wbf, wbv, name):
    """Return cached value of a single-cell defined name."""
    if name not in wbf.defined_names:
        return None
    dn = wbf.defined_names[name]
    for sheet, ref in dn.destinations:
        ref = ref.replace("$", "")
        return wbv[sheet][ref].value
    return None


def run_spot_checks(wbf, wbv):
    # SC1: ItemsPerMonth = 9890
    v = resolve_name(wbf, wbv, "ItemsPerMonth")
    check(v == 9890, "SC1 ItemsPerMonth = 9890", f"got {v!r}")

    # SC2: an Aurora monthly line ~= 1.038*730 = 757.74 (2xlarge default)
    target, found = 1.038 * 730, []
    for ws in wbv.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, (int, float)) and abs(c.value - target) < 2.0:
                    found.append(f"{ws.title}!{c.coordinate}={c.value:.2f}")
    check(
        bool(found),
        "SC2 Aurora 2xlarge monthly ≈ $757.74",
        "; ".join(found[:3]) or "no cell within $2 of 757.74",
    )

    # SC3: Opus LIKELY line ≈ $203/mo (2 calls * 9890 items, 0.8K in/0.25K out @ 5/25 per MTok)
    lo, hi, found = 150.0, 260.0, []
    items, calls = 9890, 2
    exp = items * calls * (0.8 / 1000 * 5 + 0.25 / 1000 * 25)  # ≈ 202.7
    for ws in wbv.worksheets:
        if ws.title not in ("Monthly TCO", "Overview", "Pod Build"):
            continue
        for row in ws.iter_rows():
            for c in row:
                if (
                    isinstance(c.value, (int, float))
                    and lo <= c.value <= hi
                    and abs(c.value - exp) < 15
                ):
                    found.append(f"{ws.title}!{c.coordinate}={c.value:.2f}")
    check(
        bool(found),
        f"SC3 Opus token line ≈ ${exp:.0f}/mo (code-verified LIKELY)",
        "; ".join(found[:3]) or f"no cell near {exp:.0f} on TCO sheets",
    )

    # SC4: 12mo total = sum of the 12 month columns (recompute from grid)
    ok4, detail4 = False, "12mo sheet missing"
    if "12mo Run Cost" in wbv.sheetnames:
        ws = wbv["12mo Run Cost"]
        best = None
        for row in ws.iter_rows():
            nums = [c.value for c in row if isinstance(c.value, (int, float))]
            if len(nums) >= 13:
                months, total = nums[:12], nums[-1]
                if total and abs(sum(months) - total) / max(abs(total), 1) < 0.01:
                    best = (sum(months), total)
        if best:
            ok4, detail4 = True, f"sum(months)={best[0]:.2f} vs total={best[1]:.2f}"
        else:
            detail4 = "no row where last col ≈ sum of 12 month cols"
    check(ok4, "SC4 12mo total = SUM of month columns", detail4)

    # SC5: Pod Build total points = sum of per-row points
    ok5, detail5 = False, "Pod Build sheet missing"
    if "Pod Build" in wbv.sheetnames:
        ws = wbv["Pod Build"]
        # find the Points column by header
        pts_col = None
        for row in ws.iter_rows(min_row=1, max_row=10):
            for c in row:
                if isinstance(c.value, str) and c.value.strip().lower() == "points":
                    pts_col = c.column
        if pts_col:
            vals = [
                ws.cell(row=r, column=pts_col).value for r in range(1, ws.max_row + 1)
            ]
            nums = [x for x in vals if isinstance(x, (int, float))]
            if nums:
                total, parts = max(nums), sorted(nums)[:-1]
                ok5 = abs(sum(parts) - total) < 0.01
                detail5 = f"sum(rows)={sum(parts)} vs max/total={total}"
            else:
                detail5 = "no numeric points found"
        else:
            detail5 = "no 'Points' header found"
    check(ok5, "SC5 Pod Build total points = sum of row points", detail5)


if __name__ == "__main__":
    main()
